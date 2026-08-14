"""Calibration runs — CRUD, file uploads, processing, status, downloads."""

import io
import logging
import re
import uuid
import zipfile
from datetime import UTC, date, datetime
from pathlib import Path

_log = logging.getLogger(__name__)

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from fastapi.responses import FileResponse, StreamingResponse
import unicodedata

from pydantic import BaseModel, field_validator
from sqlalchemy import case, delete, func, select, update as sql_update
from sqlalchemy.orm import Session

from ite_api.audit import write_audit
from ite_api.auth.dependencies import require_role
from ite_api.calibration.cal_loader import load_calibration_sheet, load_workbook
from ite_api.calibration.engine import RunConfig, SetpointWindow, run_one_logger
from ite_api.calibration.ref_loader import combine_refs, load_ref_auto
from ite_api.config import Settings, get_settings
from ite_api.db.models import AuditLog, User
from ite_api.db.models.calibration import (
    CalibrationRun,
    Logger,
    LoggerResult,
    RunCalibrationFile,
    RunReferenceFile,
)
from ite_api.db.session import get_session
from ite_api.storage import save_file

router = APIRouter(prefix="/api/runs", tags=["runs"])

_ALLOWED_REF_TYPES = {"text/csv", "text/plain", "application/csv"}
_ALLOWED_CAL_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
}
_REF_MAX_BYTES = 10 * 1024 * 1024
_CAL_MAX_BYTES = 50 * 1024 * 1024


# ── Schemas ────────────────────────────────────────────────────────────────

class SetpointIn(BaseModel):
    target_c: float
    start_at: datetime
    end_at: datetime


class RunCreateRequest(BaseModel):
    batch_name: str
    testing_start: datetime
    testing_end: datetime
    certificate_date: date
    threshold_c: float = 0.5
    setpoints: list[SetpointIn]
    start_cert_no: str = "0000001000"
    cert_width: int = 10
    test_date_jp: str
    doc_date_jp: str

    @field_validator("batch_name", "test_date_jp", "doc_date_jp", mode="before")
    @classmethod
    def normalize_unicode(cls, v: object) -> object:
        """NFC-normalise text fields so multi-byte characters survive the round-trip."""
        if isinstance(v, str):
            return unicodedata.normalize("NFC", v)
        return v


class RenameRunRequest(BaseModel):
    batch_name: str

    @field_validator("batch_name", mode="before")
    @classmethod
    def normalize_and_validate(cls, v: object) -> object:
        if not isinstance(v, str):
            raise ValueError("batch_name must be a string")
        v = unicodedata.normalize("NFC", v.strip())
        if not v:
            raise ValueError("batch_name cannot be empty")
        if len(v) > 200:
            raise ValueError("batch_name must be 200 characters or fewer")
        return v


class RunSummary(BaseModel):
    id: uuid.UUID
    batch_name: str
    status: str
    created_at: datetime
    completed_at: datetime | None
    logger_count: int | None = None
    pass_rate: float | None = None
    max_deviation_c: float | None = None

    model_config = {"from_attributes": True}


class RunDetail(BaseModel):
    id: uuid.UUID
    batch_name: str
    status: str
    testing_start: datetime
    testing_end: datetime
    certificate_date: date
    threshold_c: float
    setpoints: list
    start_cert_no: str
    cert_width: int
    test_date_jp: str
    doc_date_jp: str
    failure_reason: dict | None
    created_at: datetime
    completed_at: datetime | None
    reference_files: list[dict]
    calibration_file: dict | None
    results: list[dict]

    model_config = {"from_attributes": True}


class FileUploadResponse(BaseModel):
    file_id: uuid.UUID
    sha256: str
    original_name: str


class CalibrationUploadResponse(BaseModel):
    file_id: uuid.UUID
    sha256: str
    original_name: str
    sheet_names: list[str]


class StatusResponse(BaseModel):
    status: str
    message: str | None = None


# ── Helpers ────────────────────────────────────────────────────────────────

def _get_run_or_404(run_id: uuid.UUID, db: Session) -> CalibrationRun:
    run = db.get(CalibrationRun, run_id)
    if run is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="run not found")
    return run


def _settings(request: Request) -> Settings:
    return get_settings()


# ── Routes ────────────────────────────────────────────────────────────────

@router.get("", response_model=list[RunSummary])
def list_runs(
    status_filter: str | None = Query(default=None, alias="status"),
    from_date: datetime | None = Query(default=None, alias="from"),
    to_date: datetime | None = Query(default=None, alias="to"),
    q: str | None = Query(default=None),
    limit: int = Query(default=50, le=200),
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    # Single aggregation query — avoids N+1 (one SELECT per run for logger_results).
    # Only fetches verdict + max_deviation_c columns via SQL; per_setpoint JSON is never loaded.
    stats_subq = (
        select(
            LoggerResult.run_id,
            func.count(LoggerResult.id).label("logger_count"),
            (
                func.sum(case((LoggerResult.verdict == "pass", 1.0), else_=0.0))
                * 100.0
                / func.nullif(func.count(LoggerResult.id), 0)
            ).label("pass_rate"),
            func.max(LoggerResult.max_deviation_c).label("max_deviation_c"),
        )
        .group_by(LoggerResult.run_id)
        .subquery()
    )

    stmt = (
        select(
            CalibrationRun,
            stats_subq.c.logger_count,
            stats_subq.c.pass_rate,
            stats_subq.c.max_deviation_c,
        )
        .outerjoin(stats_subq, stats_subq.c.run_id == CalibrationRun.id)
        .order_by(CalibrationRun.created_at.desc())
        .limit(limit)
    )
    if status_filter:
        stmt = stmt.where(CalibrationRun.status == status_filter)
    if from_date:
        stmt = stmt.where(CalibrationRun.created_at >= from_date)
    if to_date:
        stmt = stmt.where(CalibrationRun.created_at <= to_date)
    if q:
        stmt = stmt.where(CalibrationRun.batch_name.ilike(f"%{q}%"))

    rows = db.execute(stmt).all()
    result = []
    for run, logger_count, pass_rate, max_dev in rows:
        is_complete = run.status == "complete"
        result.append(RunSummary(
            id=run.id,
            batch_name=run.batch_name,
            status=run.status,
            created_at=run.created_at,
            completed_at=run.completed_at,
            logger_count=int(logger_count) if (is_complete and logger_count is not None) else None,
            pass_rate=round(float(pass_rate), 1) if (is_complete and pass_rate is not None) else None,
            max_deviation_c=float(max_dev) if (is_complete and max_dev is not None) else None,
        ))
    return result


@router.post("", response_model=RunDetail, status_code=status.HTTP_201_CREATED)
def create_run(
    body: RunCreateRequest,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    run = CalibrationRun(
        batch_name=body.batch_name,
        status="draft",
        testing_start=body.testing_start,
        testing_end=body.testing_end,
        certificate_date=body.certificate_date,
        threshold_c=body.threshold_c,
        setpoints=[sp.model_dump(mode="json") for sp in body.setpoints],
        start_cert_no=body.start_cert_no,
        cert_width=body.cert_width,
        test_date_jp=body.test_date_jp,
        doc_date_jp=body.doc_date_jp,
        created_by=user.id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    write_audit(db, user_id=user.id, run_id=run.id, action="run.created")
    return _run_detail(run, db)


@router.get("/{run_id}", response_model=RunDetail)
def get_run(
    run_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    run = _get_run_or_404(run_id, db)
    return _run_detail(run, db)


@router.patch("/{run_id}", response_model=RunDetail)
def patch_run(
    run_id: uuid.UUID,
    body: RunCreateRequest,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    run = _get_run_or_404(run_id, db)
    if run.status != "draft":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="only draft runs can be edited")
    run.batch_name = body.batch_name
    run.testing_start = body.testing_start
    run.testing_end = body.testing_end
    run.certificate_date = body.certificate_date
    run.threshold_c = body.threshold_c
    run.setpoints = [sp.model_dump(mode="json") for sp in body.setpoints]
    run.start_cert_no = body.start_cert_no
    run.cert_width = body.cert_width
    run.test_date_jp = body.test_date_jp
    run.doc_date_jp = body.doc_date_jp
    db.commit()
    db.refresh(run)
    write_audit(db, user_id=user.id, run_id=run.id, action="run.updated")
    return _run_detail(run, db)


@router.patch("/{run_id}/rename", response_model=RunDetail)
def rename_run(
    run_id: uuid.UUID,
    body: RenameRunRequest,
    db: Session = Depends(get_session),
    user: User = require_role("admin"),
):
    """Rename a run regardless of its status (admin only)."""
    run = _get_run_or_404(run_id, db)
    old_name = run.batch_name
    run.batch_name = body.batch_name
    db.commit()
    db.refresh(run)
    write_audit(db, user_id=user.id, run_id=run.id, action="run.renamed",
                detail={"old_name": old_name, "new_name": body.batch_name})
    return _run_detail(run, db)


@router.delete("/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_run(
    run_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("admin"),
):
    run = _get_run_or_404(run_id, db)

    # Collect all on-disk paths before the DB cascade removes the records.
    ref_files = db.scalars(select(RunReferenceFile).where(RunReferenceFile.run_id == run.id)).all()
    cal_file = db.scalars(select(RunCalibrationFile).where(RunCalibrationFile.run_id == run.id)).first()
    results = db.scalars(select(LoggerResult).where(LoggerResult.run_id == run.id)).all()

    write_audit(db, user_id=user.id, run_id=run.id, action="run.deleted")
    db.delete(run)
    db.commit()

    # Remove files from the data volume after the DB transaction commits.
    paths_to_remove: list[Path] = []
    for f in ref_files:
        paths_to_remove.append(Path(f.stored_path))
    if cal_file:
        paths_to_remove.append(Path(cal_file.stored_path))
    for r in results:
        if r.cert_path:
            paths_to_remove.append(Path(r.cert_path))

    for p in paths_to_remove:
        try:
            p.unlink(missing_ok=True)
        except OSError as e:
            _log.warning("Could not delete file %s: %s", p, e)

    # Best-effort: remove the now-empty run directory.
    settings = get_settings()
    run_dir = settings.data_dir / "runs" / str(run_id)
    try:
        if run_dir.exists():
            import shutil
            shutil.rmtree(run_dir, ignore_errors=True)
    except OSError as e:
        _log.warning("Could not remove run directory %s: %s", run_dir, e)


# ── File uploads ──────────────────────────────────────────────────────────

@router.post("/{run_id}/references", response_model=FileUploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_reference(
    run_id: uuid.UUID,
    file: UploadFile,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
    request: Request = None,
):
    run = _get_run_or_404(run_id, db)
    if run.status != "draft":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="run is not in draft state")
    data = await file.read()
    if len(data) > _REF_MAX_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="reference file too large (max 10 MB)")
    settings = get_settings()
    stored, sha256 = save_file(
        data, run_id=run_id, sub="references",
        original_name=file.filename or "reference.csv", data_dir=settings.data_dir
    )
    ref_file = RunReferenceFile(
        run_id=run_id,
        original_name=file.filename or "reference.csv",
        stored_path=str(stored),
        sha256=sha256,
    )
    db.add(ref_file)
    db.commit()
    db.refresh(ref_file)
    return FileUploadResponse(file_id=ref_file.id, sha256=sha256, original_name=ref_file.original_name)


@router.post("/{run_id}/calibration", response_model=CalibrationUploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_calibration(
    run_id: uuid.UUID,
    file: UploadFile,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    run = _get_run_or_404(run_id, db)
    if run.status != "draft":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="run is not in draft state")
    data = await file.read()
    if len(data) > _CAL_MAX_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="calibration file too large (max 50 MB)")
    settings = get_settings()
    stored, sha256 = save_file(
        data, run_id=run_id, sub="calibration",
        original_name=file.filename or "workbook.xlsx", data_dir=settings.data_dir
    )
    # Detect sheet names
    import openpyxl
    wb_obj = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_names = wb_obj.sheetnames
    wb_obj.close()

    # Replace any existing calibration file record
    existing = db.scalars(select(RunCalibrationFile).where(RunCalibrationFile.run_id == run_id)).first()
    if existing:
        db.delete(existing)
        db.flush()

    cal_file = RunCalibrationFile(
        run_id=run_id,
        original_name=file.filename or "workbook.xlsx",
        stored_path=str(stored),
        sha256=sha256,
        sheet_names=sheet_names,
    )
    db.add(cal_file)
    db.commit()
    db.refresh(cal_file)
    return CalibrationUploadResponse(
        file_id=cal_file.id, sha256=sha256,
        original_name=cal_file.original_name, sheet_names=sheet_names
    )


@router.delete("/{run_id}/files/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_file(
    run_id: uuid.UUID,
    file_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    _get_run_or_404(run_id, db)
    ref = db.get(RunReferenceFile, file_id)
    if ref and ref.run_id == run_id:
        db.delete(ref)
        db.commit()
        return
    cal = db.get(RunCalibrationFile, file_id)
    if cal and cal.run_id == run_id:
        db.delete(cal)
        db.commit()
        return
    raise HTTPException(status.HTTP_404_NOT_FOUND, detail="file not found")


# ── Processing ────────────────────────────────────────────────────────────

@router.post("/{run_id}/process", status_code=status.HTTP_202_ACCEPTED)
def process_run(
    run_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    _get_run_or_404(run_id, db)

    ref_files = db.scalars(select(RunReferenceFile).where(RunReferenceFile.run_id == run_id)).all()
    cal_file = db.scalars(select(RunCalibrationFile).where(RunCalibrationFile.run_id == run_id)).first()

    if not ref_files:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="no reference files uploaded")
    if cal_file is None:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="no calibration workbook uploaded")

    # Atomic transition: only one concurrent request can win this UPDATE.
    # If two requests race, only one gets rowcount=1; the other hits the 409.
    result = db.execute(
        sql_update(CalibrationRun)
        .where(CalibrationRun.id == run_id)
        .where(CalibrationRun.status.in_(("draft", "failed")))
        .values(status="processing")
        .returning(CalibrationRun.id)
    )
    db.commit()
    if not result.fetchone():
        run = db.get(CalibrationRun, run_id)
        current = run.status if run else "unknown"
        raise HTTPException(status.HTTP_409_CONFLICT, detail=f"run status is '{current}', cannot process")

    write_audit(db, user_id=user.id, run_id=run_id, action="run.processing_started")

    settings = get_settings()
    background_tasks.add_task(
        _run_processing_task,
        run_id=run_id,
        ref_paths=[Path(f.stored_path) for f in ref_files],
        cal_path=Path(cal_file.stored_path),
        settings=settings,
    )
    return {"job_id": str(run_id)}


@router.get("/{run_id}/status", response_model=StatusResponse)
def get_status(
    run_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    run = _get_run_or_404(run_id, db)
    msg = None
    if run.status == "failed" and run.failure_reason:
        msg = run.failure_reason.get("message")
    return StatusResponse(status=run.status, message=msg)


# ── Results & certificates ────────────────────────────────────────────────

@router.get("/{run_id}/results")
def list_results(
    run_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    _get_run_or_404(run_id, db)
    results = db.scalars(select(LoggerResult).where(LoggerResult.run_id == run_id)).all()
    return [
        {
            "id": str(r.id),
            "sheet_name": r.sheet_name,
            "verdict": r.verdict,
            "max_deviation_c": float(r.max_deviation_c) if r.max_deviation_c is not None else None,
            "cert_no": r.cert_no,
            "per_setpoint": r.per_setpoint,
        }
        for r in results
    ]


@router.get("/{run_id}/results/{result_id}/certificate")
def download_certificate(
    run_id: uuid.UUID,
    result_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    _get_run_or_404(run_id, db)
    result = db.get(LoggerResult, result_id)
    if result is None or result.run_id != run_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="result not found")
    if not result.cert_path or not Path(result.cert_path).exists():
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="certificate file not found")
    write_audit(db, user_id=user.id, run_id=run_id, action="cert.downloaded",
                detail={"result_id": str(result_id), "cert_no": result.cert_no})
    filename = Path(result.cert_path).name
    return FileResponse(result.cert_path, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/{run_id}/audit")
def get_audit(
    run_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    _get_run_or_404(run_id, db)
    entries = db.scalars(
        select(AuditLog).where(AuditLog.run_id == run_id).order_by(AuditLog.at.desc())
    ).all()
    user_cache: dict[uuid.UUID, User] = {}
    result = []
    for e in entries:
        actor = None
        if e.user_id:
            if e.user_id not in user_cache:
                user_cache[e.user_id] = db.get(User, e.user_id)
            u = user_cache[e.user_id]
            if u:
                actor = {"id": str(u.id), "full_name": u.full_name, "email": u.email}
        result.append({"action": e.action, "at": e.at.isoformat(), "detail": e.detail, "user": actor})
    return result


class PatchDatesRequest(BaseModel):
    test_date_jp: str | None = None
    doc_date_jp: str | None = None


@router.patch("/{run_id}/dates", response_model=RunDetail)
def patch_dates(
    run_id: uuid.UUID,
    body: PatchDatesRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    """Update test date and/or cert date on a completed run, then regenerate all certificates."""
    run = _get_run_or_404(run_id, db)
    if run.status != "complete":
        raise HTTPException(status.HTTP_409_CONFLICT, detail="only complete runs can have dates edited")
    if not body.test_date_jp and not body.doc_date_jp:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="no date provided")

    changes: dict = {}
    if body.test_date_jp and body.test_date_jp != run.test_date_jp:
        changes["test_date_jp"] = {"old": run.test_date_jp, "new": body.test_date_jp}
        run.test_date_jp = body.test_date_jp
    if body.doc_date_jp and body.doc_date_jp != run.doc_date_jp:
        changes["doc_date_jp"] = {"old": run.doc_date_jp, "new": body.doc_date_jp}
        run.doc_date_jp = body.doc_date_jp

    if not changes:
        return _run_detail(run, db)

    db.commit()
    db.refresh(run)
    write_audit(db, user_id=user.id, run_id=run.id, action="run.dates_edited", detail=changes)

    settings = get_settings()
    background_tasks.add_task(_regenerate_all_certificates, run.id, settings)
    return _run_detail(run, db)


def _regenerate_all_certificates(run_id: uuid.UUID, settings) -> None:
    from ite_api.db.session import _SessionLocal
    assert _SessionLocal is not None
    with _SessionLocal() as db:
        run = db.get(CalibrationRun, run_id)
        if run is None:
            return
        results = db.scalars(select(LoggerResult).where(LoggerResult.run_id == run_id)).all()
        for result in results:
            try:
                _regenerate_certificate(run, result, settings, db)
            except Exception:
                _log.exception("Failed to regenerate cert for result %s during date update", result.id)
        db.commit()


@router.get("/{run_id}/results.zip")
def download_all_certs(
    run_id: uuid.UUID,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    run = _get_run_or_404(run_id, db)
    results = db.scalars(select(LoggerResult).where(LoggerResult.run_id == run_id)).all()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            if r.cert_path and Path(r.cert_path).exists():
                zf.write(r.cert_path, Path(r.cert_path).name)
    buf.seek(0)
    safe_name = re.sub(r"[^\w\-]", "_", run.batch_name)[:50]
    write_audit(db, user_id=user.id, run_id=run_id, action="certs.zip_downloaded")
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_certificates.zip"'},
    )


@router.get("/by-cert-no/{cert_no}")
def find_by_cert_no(
    cert_no: str,
    db: Session = Depends(get_session),
    user: User = require_role("viewer"),
):
    result = db.scalars(select(LoggerResult).where(LoggerResult.cert_no == cert_no)).first()
    if result is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="certificate not found")
    run = db.get(CalibrationRun, result.run_id)
    return {
        "result_id": str(result.id),
        "run_id": str(result.run_id),
        "cert_no": result.cert_no,
        "sheet_name": result.sheet_name,
        "verdict": result.verdict,
        "batch_name": run.batch_name if run else None,
        "certificate_date": run.certificate_date.isoformat() if run and run.certificate_date else None,
    }


# ── Background task ───────────────────────────────────────────────────────

def _run_processing_task(
    *,
    run_id: uuid.UUID,
    ref_paths: list[Path],
    cal_path: Path,
    settings: Settings,
) -> None:
    from ite_api.db.session import _SessionLocal
    assert _SessionLocal is not None
    with _SessionLocal() as db:
        run = db.get(CalibrationRun, run_id)
        if run is None:
            return
        try:
            _do_process(run, ref_paths, cal_path, settings, db)
        except Exception as exc:  # noqa: BLE001
            import traceback
            _log.exception("Processing failed for run %s", run_id)
            run.status = "failed"
            run.failure_reason = {
                "message": f"{type(exc).__name__}: {exc}",
                "traceback": traceback.format_exc(),
            }
            db.commit()


def _do_process(
    run: CalibrationRun,
    ref_paths: list[Path],
    cal_path: Path,
    settings: Settings,
    db: Session,
) -> None:
    from ite_api.calibration.matcher import find_values_for_target

    # Wipe any results from a previous partial/failed attempt so retries are clean.
    db.execute(delete(LoggerResult).where(LoggerResult.run_id == run.id))
    db.flush()

    ref_df = combine_refs([load_ref_auto(p) for p in ref_paths])
    wb, sheet_names = load_workbook(cal_path)

    template_path = Path(run.template_path) if run.template_path else _default_template()

    setpoints = [
        SetpointWindow(
            target=sp["target_c"],
            start=datetime.fromisoformat(sp["start_at"]) if isinstance(sp["start_at"], str) else sp["start_at"],
            end=datetime.fromisoformat(sp["end_at"]) if isinstance(sp["end_at"], str) else sp["end_at"],
        )
        for sp in run.setpoints
    ]

    start = int(run.start_cert_no)
    threshold = float(run.threshold_c)

    for idx, name in enumerate(sheet_names):
        cert_no = str(start + idx).zfill(run.cert_width)
        run_cfg = RunConfig(
            cert_no=cert_no,
            serial=name.strip(),
            test_date_jp=run.test_date_jp,
            doc_date_jp=run.doc_date_jp,
            template_path=template_path,
            output_dir=settings.data_dir / "runs" / str(run.id) / "certificates",
            setpoints=setpoints,
        )
        out_path = run_one_logger(run_cfg, sheet_name=name, wb=wb, ref_df=ref_df)

        logger = db.scalars(select(Logger).where(Logger.serial_no == name.strip())).first()
        if logger is None:
            logger = Logger(serial_no=name.strip())
            db.add(logger)
            db.flush()

        cal_df = load_calibration_sheet(wb, name)
        per_sp = []
        deviations = []
        for sp in setpoints:
            ref_v, cal_v, _ = find_values_for_target(cal_df, ref_df, sp.target, sp.start, sp.end)
            dev = abs(ref_v - cal_v) if ref_v is not None and cal_v is not None else None
            per_sp.append({
                "target_c": sp.target,
                "ref_c": ref_v,
                "cal_c": cal_v,
                "dev_c": round(dev, 3) if dev is not None else None,
                "within_tol": (dev is not None and dev <= threshold),
            })
            if dev is not None:
                deviations.append(dev)

        max_dev = max(deviations) if deviations else None
        verdict = "pass" if (max_dev is not None and max_dev <= threshold) else "fail"

        lr = LoggerResult(
            run_id=run.id,
            logger_id=logger.id,
            sheet_name=name,
            verdict=verdict,
            max_deviation_c=max_dev,
            per_setpoint=per_sp,
            cert_no=cert_no,
            cert_path=str(out_path),
        )
        db.add(lr)

    run.status = "complete"
    run.completed_at = datetime.now(UTC)
    run.failure_reason = None
    db.commit()


def _default_template() -> Path:
    """Fall back to the fixture template if no template_path is set on the run."""
    here = Path(__file__).parent.parent
    candidate = here / "calibration" / "template.docx"
    if candidate.exists():
        return candidate
    raise RuntimeError("No template configured and no default template found")


# ── Detail builder ────────────────────────────────────────────────────────

# ── Manual deviation correction ───────────────────────────────────────────

class DeviationCorrection(BaseModel):
    setpoint_index: int
    new_deviation: float


@router.patch("/{run_id}/results/{result_id}", status_code=status.HTTP_200_OK)
def correct_deviation(
    run_id: uuid.UUID,
    result_id: uuid.UUID,
    body: DeviationCorrection,
    db: Session = Depends(get_session),
    user: User = require_role("engineer"),
):
    run = _get_run_or_404(run_id, db)
    result = db.get(LoggerResult, result_id)
    if result is None or result.run_id != run_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="result not found")

    per_sp = list(result.per_setpoint)
    if body.setpoint_index < 0 or body.setpoint_index >= len(per_sp):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="invalid setpoint_index")

    sp = dict(per_sp[body.setpoint_index])
    ref_c = sp.get("ref_c")
    if ref_c is None:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="no reference value for this setpoint")

    new_dev = round(abs(body.new_deviation), 3)
    # Preserve sign direction of original deviation; default to ref - cal convention
    original_dev = sp.get("dev_c")
    if original_dev is not None and sp.get("cal_c") is not None and sp.get("ref_c") is not None:
        sign = 1 if sp["cal_c"] <= sp["ref_c"] else -1
    else:
        sign = 1
    new_cal_c = round(ref_c - sign * new_dev, 3)

    threshold = float(run.threshold_c)
    if "original_dev_c" not in sp:
        sp["original_dev_c"] = sp.get("dev_c")
    sp["dev_c"] = new_dev
    sp["cal_c"] = new_cal_c
    sp["within_tol"] = new_dev <= threshold
    sp["manually_corrected"] = True
    per_sp[body.setpoint_index] = sp

    devs = [s["dev_c"] for s in per_sp if s.get("dev_c") is not None]
    new_max = max(devs) if devs else None
    new_verdict = "pass" if (new_max is not None and new_max <= threshold) else "fail"

    result.per_setpoint = per_sp
    result.max_deviation_c = new_max
    result.verdict = new_verdict

    # Patch the Word certificate
    settings = get_settings()
    cert_path = Path(result.cert_path) if result.cert_path else None
    if cert_path and cert_path.exists():
        try:
            from docx import Document
            from ite_api.calibration.docx_filler import fill_results_table
            doc = Document(str(cert_path))
            ordered = [(s.get("ref_c"), s.get("cal_c")) for s in per_sp]
            fill_results_table(doc, ordered)
            doc.save(str(cert_path))
        except Exception:
            _log.exception("Patching docx failed for result %s — regenerating", result_id)
            _regenerate_certificate(run, result, settings, db)
    else:
        _regenerate_certificate(run, result, settings, db)

    write_audit(db, user_id=user.id, run_id=run_id, action="result.deviation_corrected", detail={
        "result_id": str(result_id),
        "sheet_name": result.sheet_name,
        "setpoint_index": body.setpoint_index,
        "target_c": sp.get("target_c"),
        "old_dev_c": sp.get("dev_c"),
        "new_dev_c": new_dev,
        "new_verdict": new_verdict,
    })
    db.commit()
    return {
        "id": str(result.id),
        "sheet_name": result.sheet_name,
        "verdict": result.verdict,
        "max_deviation_c": float(result.max_deviation_c) if result.max_deviation_c is not None else None,
        "cert_no": result.cert_no,
        "per_setpoint": result.per_setpoint,
    }



def _regenerate_certificate(
    run: CalibrationRun,
    result: LoggerResult,
    settings,
    db: Session,
) -> None:
    from ite_api.calibration.engine import RunConfig, SetpointWindow, run_one_logger
    from ite_api.calibration.cal_loader import load_workbook
    from ite_api.calibration.ref_loader import combine_refs, load_ref_auto

    ref_files = db.scalars(select(RunReferenceFile).where(RunReferenceFile.run_id == run.id)).all()
    cal_file = db.scalars(select(RunCalibrationFile).where(RunCalibrationFile.run_id == run.id)).first()
    if not ref_files or cal_file is None:
        raise RuntimeError("Cannot regenerate: missing uploaded files")

    ref_df = combine_refs([load_ref_auto(Path(f.stored_path)) for f in ref_files])
    wb, _ = load_workbook(Path(cal_file.stored_path))

    template_path = Path(run.template_path) if run.template_path else _default_template()
    setpoints = [
        SetpointWindow(
            target=sp["target_c"],
            start=datetime.fromisoformat(sp["start_at"]) if isinstance(sp["start_at"], str) else sp["start_at"],
            end=datetime.fromisoformat(sp["end_at"]) if isinstance(sp["end_at"], str) else sp["end_at"],
        )
        for sp in run.setpoints
    ]

    run_cfg = RunConfig(
        cert_no=result.cert_no or "",
        serial=result.sheet_name.strip(),
        test_date_jp=run.test_date_jp,
        doc_date_jp=run.doc_date_jp,
        template_path=template_path,
        output_dir=settings.data_dir / "runs" / str(run.id) / "certificates",
        setpoints=setpoints,
    )
    out_path = run_one_logger(run_cfg, sheet_name=result.sheet_name, wb=wb, ref_df=ref_df)
    result.cert_path = str(out_path)


def _run_detail(run: CalibrationRun, db: Session) -> RunDetail:
    ref_files = db.scalars(select(RunReferenceFile).where(RunReferenceFile.run_id == run.id)).all()
    cal_file = db.scalars(select(RunCalibrationFile).where(RunCalibrationFile.run_id == run.id)).first()
    results = db.scalars(select(LoggerResult).where(LoggerResult.run_id == run.id)).all()
    return RunDetail(
        id=run.id,
        batch_name=run.batch_name,
        status=run.status,
        testing_start=run.testing_start,
        testing_end=run.testing_end,
        certificate_date=run.certificate_date,
        threshold_c=float(run.threshold_c),
        setpoints=run.setpoints,
        start_cert_no=run.start_cert_no,
        cert_width=run.cert_width,
        test_date_jp=run.test_date_jp,
        doc_date_jp=run.doc_date_jp,
        failure_reason=run.failure_reason,
        created_at=run.created_at,
        completed_at=run.completed_at,
        reference_files=[
            {"id": str(f.id), "original_name": f.original_name, "sha256": f.sha256}
            for f in ref_files
        ],
        calibration_file={
            "id": str(cal_file.id),
            "original_name": cal_file.original_name,
            "sha256": cal_file.sha256,
            "sheet_names": cal_file.sheet_names,
        } if cal_file else None,
        results=[
            {
                "id": str(r.id),
                "sheet_name": r.sheet_name,
                "verdict": r.verdict,
                "max_deviation_c": float(r.max_deviation_c) if r.max_deviation_c is not None else None,
                "cert_no": r.cert_no,
                "per_setpoint": r.per_setpoint,
            }
            for r in results
        ],
    )


