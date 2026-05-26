"""Calibration logger XLSX loading — one sheet per logger."""

from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


def load_workbook(path: Path) -> tuple[openpyxl.workbook.workbook.Workbook, list[str]]:
    """Open the calibration workbook in data-only mode; return (workbook, sheet_names)."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    return wb, wb.sheetnames


def load_calibration_sheet(wb, sheet_name: str) -> pd.DataFrame:
    """Load one logger sheet into DataFrame with columns ['timestamp', 'temp']."""
    ws = wb[sheet_name]
    rows: list[tuple[datetime, float]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 4 or row[1] is None:
            continue
        ts_value = row[1]
        if isinstance(ts_value, datetime):
            ts = ts_value
        else:
            try:
                ts = datetime.strptime(str(ts_value).strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        try:
            temp = float(str(row[3]).strip())
        except (ValueError, TypeError):
            continue
        rows.append((ts, temp))
    df = pd.DataFrame(rows, columns=["timestamp", "temp"])
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    df["temp"] = df["temp"].astype(float)
    return df
