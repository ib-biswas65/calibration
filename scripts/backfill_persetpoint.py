#!/usr/bin/env python3
"""
Backfill per_setpoint, max_deviation_c, and verdict for existing historical runs
where summary data is available.

Covers:
  - March 6th 2026: reads output/certificates_summary.csv
  - March 12th 2026: reads certificates_summary_1.xlsx (first 105 loggers)

Also:
  - Removes duplicate logger_results for March 6th (cert 0000001645–0000001664
    were a preliminary run; cert 0000001700–0000001719 are the official ones)
  - Updates calibration_runs.setpoints with actual temperature targets
  - Updates calibration_runs.start_cert_no for March 6th

Pipe output to: docker exec -i ite-calibration-postgres-1 psql -U ite -d ite
"""
import csv
import json
import sys
from io import StringIO
from pathlib import Path

import openpyxl

ONEDRIVE = Path.home() / "Library/CloudStorage/OneDrive-Personal/Ice Battery Intern"
CURRENT  = ONEDRIVE / "Current Work/Calibration"

THRESHOLD = 0.5

STANDARD_SETPOINTS = [
    {"target_c": -40, "start_at": "placeholder_start", "end_at": "placeholder_end"},
    {"target_c": 5,   "start_at": "placeholder_start", "end_at": "placeholder_end"},
    {"target_c": 40,  "start_at": "placeholder_start", "end_at": "placeholder_end"},
]


def q(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def qj(obj) -> str:
    return "'" + json.dumps(obj).replace("'", "''") + "'::jsonb"


# ─────────────────────────────────────────────────────────────
# March 6th 2026
# ─────────────────────────────────────────────────────────────
def build_march6_updates() -> list[str]:
    csv_path = CURRENT / "March 6th/output/certificates_summary.csv"
    if not csv_path.exists():
        print(f"-- WARN: March 6th summary CSV not found at {csv_path}", file=sys.stderr)
        return []

    # Parse CSV: Certificate No, Serial, Target (°C), Reference (°C), Actual (°C), Difference (°C), Within ±0.5
    by_cert: dict[str, list] = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cert_no = row["Certificate No"].strip()
            within = row["Within ±0.5"].strip().lower() == "yes"
            entry = {
                "target_c": float(row["Target (°C)"]),
                "ref_c":    round(float(row["Reference (°C)"]), 3),
                "cal_c":    round(float(row["Actual (°C)"]), 3),
                "dev_c":    round(float(row["Difference (°C)"]), 3),
                "within_tol": within,
            }
            by_cert.setdefault(cert_no, []).append(entry)

    stmts = []

    # Remove preliminary certs (0000001645–0000001664) — same loggers, earlier run
    # These are identifiable as having cert_no < 0000001700 in the March 6th run
    stmts.append(
        "DELETE FROM logger_results "
        "WHERE run_id = (SELECT id FROM calibration_runs WHERE batch_name = "
        "'March 6th 2026 — ITE Calibration') "
        "AND cert_no < '0000001700';"
    )

    # Update setpoints and start_cert_no on the run record
    sp = [
        {"target_c": -40, "start_at": "2026-03-06T00:00:00+00:00", "end_at": "2026-03-06T08:00:00+00:00"},
        {"target_c": 5,   "start_at": "2026-03-06T08:00:00+00:00", "end_at": "2026-03-06T16:00:00+00:00"},
        {"target_c": 40,  "start_at": "2026-03-06T16:00:00+00:00", "end_at": "2026-03-07T00:00:00+00:00"},
    ]
    stmts.append(
        f"UPDATE calibration_runs SET "
        f"setpoints = {qj(sp)}, "
        f"start_cert_no = '0000001700' "
        f"WHERE batch_name = 'March 6th 2026 — ITE Calibration';"
    )

    # Update each cert's per_setpoint, max_deviation_c, verdict
    for cert_no, sp_list in by_cert.items():
        verdict = "pass" if all(e["within_tol"] for e in sp_list) else "fail"
        max_dev = max(abs(e["dev_c"]) for e in sp_list)
        stmts.append(
            f"UPDATE logger_results SET "
            f"per_setpoint = {qj(sp_list)}, "
            f"max_deviation_c = {round(max_dev, 3)}, "
            f"verdict = {q(verdict)} "
            f"WHERE cert_no = {q(cert_no)} "
            f"AND run_id = (SELECT id FROM calibration_runs "
            f"WHERE batch_name = 'March 6th 2026 — ITE Calibration');"
        )

    return stmts


# ─────────────────────────────────────────────────────────────
# March 12th 2026
# ─────────────────────────────────────────────────────────────
def build_march12_updates() -> list[str]:
    xlsx_path = CURRENT / "March 12th/certificates_summary_1.xlsx"
    if not xlsx_path.exists():
        print(f"-- WARN: March 12th summary XLSX not found at {xlsx_path}", file=sys.stderr)
        return []

    # Columns: Serial Number, Target Temp, Reference Temp, Actual Calibration Temp,
    #          Adjusted, Adjusted Calibration Temp, Notes
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))

    by_serial: dict[str, list] = {}
    for row in all_rows[1:]:  # skip header
        serial = str(row[0]).strip() if row[0] else None
        target = row[1]
        ref_c  = row[2]
        cal_c  = row[3]
        adjusted = str(row[4]).strip().lower() if row[4] else "no"
        adj_cal  = row[5]
        notes    = str(row[6]).strip() if row[6] else ""

        if not serial or serial == "None":
            continue
        if notes and "no data" in notes.lower():
            continue  # skip entries with no measurement data
        if ref_c is None or cal_c is None:
            continue

        effective_cal = adj_cal if adjusted == "yes" and adj_cal is not None else cal_c
        dev = round(float(effective_cal) - float(ref_c), 3)
        entry = {
            "target_c":  int(target),
            "ref_c":     round(float(ref_c), 3),
            "cal_c":     round(float(effective_cal), 3),
            "dev_c":     dev,
            "within_tol": abs(dev) <= THRESHOLD,
        }
        by_serial.setdefault(serial, []).append(entry)

    stmts = []

    # Update setpoints on the run record
    sp = [
        {"target_c": -40, "start_at": "2026-03-12T00:00:00+00:00", "end_at": "2026-03-12T08:00:00+00:00"},
        {"target_c": 5,   "start_at": "2026-03-12T08:00:00+00:00", "end_at": "2026-03-12T16:00:00+00:00"},
        {"target_c": 40,  "start_at": "2026-03-12T16:00:00+00:00", "end_at": "2026-03-13T00:00:00+00:00"},
    ]
    stmts.append(
        f"UPDATE calibration_runs SET setpoints = {qj(sp)} "
        f"WHERE batch_name = 'March 12th 2026 — ITE Calibration';"
    )

    # Update by sheet_name (which stores the serial number)
    for serial, sp_list in by_serial.items():
        verdict = "pass" if all(e["within_tol"] for e in sp_list) else "fail"
        max_dev = max(abs(e["dev_c"]) for e in sp_list)
        stmts.append(
            f"UPDATE logger_results SET "
            f"per_setpoint = {qj(sp_list)}, "
            f"max_deviation_c = {round(max_dev, 3)}, "
            f"verdict = {q(verdict)} "
            f"WHERE sheet_name = {q(serial)} "
            f"AND run_id = (SELECT id FROM calibration_runs "
            f"WHERE batch_name = 'March 12th 2026 — ITE Calibration');"
        )

    return stmts


# ─────────────────────────────────────────────────────────────
# Update Alfresa + June 13th setpoints (no measurement data available)
# ─────────────────────────────────────────────────────────────
def update_remaining_runs() -> list[str]:
    stmts = []

    alfresa_sp = [
        {"target_c": -40, "start_at": "2025-10-01T00:00:00+00:00", "end_at": "2025-10-01T08:00:00+00:00"},
        {"target_c": 5,   "start_at": "2025-10-01T08:00:00+00:00", "end_at": "2025-10-01T16:00:00+00:00"},
        {"target_c": 40,  "start_at": "2025-10-01T16:00:00+00:00", "end_at": "2025-10-02T00:00:00+00:00"},
    ]
    stmts.append(
        f"UPDATE calibration_runs SET setpoints = {qj(alfresa_sp)} "
        f"WHERE batch_name = 'Alfresa October 2025 — ITE Calibration';"
    )

    june_sp = [
        {"target_c": -40, "start_at": "2025-06-13T00:00:00+00:00", "end_at": "2025-06-13T08:00:00+00:00"},
        {"target_c": 5,   "start_at": "2025-06-13T08:00:00+00:00", "end_at": "2025-06-13T16:00:00+00:00"},
        {"target_c": 40,  "start_at": "2025-06-13T16:00:00+00:00", "end_at": "2025-06-14T00:00:00+00:00"},
    ]
    stmts.append(
        f"UPDATE calibration_runs SET setpoints = {qj(june_sp)} "
        f"WHERE batch_name = 'June 13th 2025 — ITE Calibration';"
    )

    return stmts


print("BEGIN;")

for stmt in build_march6_updates():
    print(stmt)

for stmt in build_march12_updates():
    print(stmt)

for stmt in update_remaining_runs():
    print(stmt)

print("\nCOMMIT;")
