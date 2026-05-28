#!/usr/bin/env python3
"""
Generate SQL INSERT statements for Old Work calibration batches (Oct 2025, Nov 2025).
These batches have measurement data in xlsx reports but no individual certificate files.

Pipe output to: docker exec -i ite-calibration-postgres-1 psql -U ite -d ite
"""
import json
import statistics
import sys
import uuid
from pathlib import Path

import openpyxl

ONEDRIVE = Path.home() / "Library/CloudStorage/OneDrive-Personal/Ice Battery Intern"
OLD_WORK = ONEDRIVE / "Old Work/Caliberation"

THRESHOLD = 0.5


def q(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def qj(obj) -> str:
    return "'" + json.dumps(obj).replace("'", "''") + "'::jsonb"


def col_means(rows: list, col_idx: int) -> float | None:
    vals = [r[col_idx] for r in rows if r[col_idx] is not None and isinstance(r[col_idx], (int, float))]
    return statistics.mean(vals) if vals else None


def compute_persetpoint_from_wide_report(path: Path, target_c: int, ref_col_range: tuple, first_serial_col: int) -> dict:
    """
    Read a wide-format xlsx report (rows=time, cols=loggers) and compute per-logger stats.
    Returns {serial_no: {"ref_c": float, "cal_c": float, "dev_c": float, "within_tol": bool}}

    Per-logger stabilization: only rows where that logger's own reading is within ±5°C of
    target_c are used. This avoids warm-up rows skewing averages when loggers were loaded
    into the chamber at different times.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[0]
    data = all_rows[1:]

    # Stable-row filter window for per-logger readings
    stab_lo, stab_hi = target_c - 5.0, target_c + 5.0

    # Reference mean: only from rows where refs themselves are in the stable window
    ref_start, ref_end = ref_col_range
    ref_vals = []
    for row in data:
        row_refs = [row[ci] for ci in range(ref_start, ref_end + 1)
                    if ci < len(row) and isinstance(row[ci], (int, float))]
        if row_refs and all(stab_lo <= v <= stab_hi for v in row_refs):
            ref_vals.extend(row_refs)
    avg_ref = statistics.mean(ref_vals) if ref_vals else float(target_c)

    # Compute mean per logger using only stabilized rows for that logger
    results = {}
    for ci in range(first_serial_col, len(header)):
        serial = header[ci]
        if serial is None or not isinstance(serial, (str, int)):
            continue
        serial = str(serial).strip()
        if not serial:
            continue
        cal_vals = [r[ci] for r in data
                    if ci < len(r) and r[ci] is not None and isinstance(r[ci], (int, float))
                    and stab_lo <= r[ci] <= stab_hi]
        if not cal_vals:
            continue
        cal_mean = statistics.mean(cal_vals)
        dev = round(cal_mean - avg_ref, 3)
        results[serial] = {
            "ref_c": round(avg_ref, 3),
            "cal_c": round(cal_mean, 3),
            "dev_c": dev,
            "within_tol": abs(dev) <= THRESHOLD,
        }
    return results


# ─────────────────────────────────────────────────────────────
# October 2025 batch
# report_75.xlsx: wide format, one sheet, 3 temperature segments
# columns: Date Time, temp (mean of all refs), ref_1, ref_2, ref_3, ref_4, serial1...
# NOTE: ref_1 and ref_4 are OUTSIDE the test chamber; ref_2 and ref_3 are INSIDE.
# Segmentation is done by ref_2/ref_3 (inside-chamber), not the "temp" average.
# Setpoints are +40°C, +5°C, -40°C (the "temp" column appears as 41, 12, -10
# because it averages inside+outside loggers).
# ─────────────────────────────────────────────────────────────
def build_oct_persetpoint() -> dict:
    path = OLD_WORK / "Calibration Oct/Report/report_75.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[0]
    data = all_rows[1:]

    # col 0=DateTime, 1=temp(mean), 2=ref_1, 3=ref_2, 4=ref_3, 5=ref_4, 6+=serials
    # ref_2 (col 3) and ref_3 (col 4) are inside the chamber — use them for segmentation
    serial_cols = [(ci, str(header[ci]).strip()) for ci in range(6, len(header))
                   if header[ci] is not None and str(header[ci]).strip().startswith("19")]

    # segment by inside-chamber ref cols (3 and 4)
    segments = [
        (40,  35.0,  45.0),   # all 4 refs are inside hot chamber
        (5,    3.0,   8.0),   # only ref_2/ref_3 inside +5°C chamber
        (-40, -45.0, -35.0),  # only ref_2/ref_3 inside -40°C chamber
    ]
    # ref columns to average per segment (cols 2-5 for +40; cols 3-4 only for +5/-40)
    seg_ref_cols = {40: [2, 3, 4, 5], 5: [3, 4], -40: [3, 4]}

    all_serial_sp: dict[str, list] = {}

    for target_c, low, high in segments:
        ref_ci_list = seg_ref_cols[target_c]
        seg_rows = [r for r in data
                    if r[3] is not None and r[4] is not None
                    and isinstance(r[3], (int, float)) and isinstance(r[4], (int, float))
                    and low <= r[3] <= high and low <= r[4] <= high]
        if not seg_rows:
            print(f"-- WARN: no Oct rows for target {target_c}°C", file=sys.stderr)
            continue

        ref_vals = [r[ci] for r in seg_rows for ci in ref_ci_list
                    if r[ci] is not None and isinstance(r[ci], (int, float))]
        avg_ref = statistics.mean(ref_vals) if ref_vals else float(target_c)

        for ci, serial in serial_cols:
            cal_vals = [r[ci] for r in seg_rows if r[ci] is not None and isinstance(r[ci], (int, float))]
            if not cal_vals:
                continue
            cal_mean = statistics.mean(cal_vals)
            dev = round(cal_mean - avg_ref, 3)
            entry = {
                "target_c": target_c,
                "ref_c": round(avg_ref, 3),
                "cal_c": round(cal_mean, 3),
                "dev_c": dev,
                "within_tol": abs(dev) <= THRESHOLD,
            }
            all_serial_sp.setdefault(serial, []).append(entry)

    return all_serial_sp


# ─────────────────────────────────────────────────────────────
# November 2025 batch
# Three separate files per temperature setpoint
# ─────────────────────────────────────────────────────────────
def build_nov_persetpoint() -> dict:
    reports = [
        # (file, target_c, first_serial_col, ref_col_range)
        (OLD_WORK / "Calibration November/Report/report_minus40_A.xlsx", -40, 4, (1, 2)),
        (OLD_WORK / "Calibration November/Report/report_plus5_A.xlsx",   5,  6, (1, 4)),
        (OLD_WORK / "Calibration November/Report/report_plus40.xlsx",    40, 3, (1, 2)),
    ]

    all_serial_sp: dict[str, list] = {}

    for path, target_c, first_serial_col, (ref_start, ref_end) in reports:
        sp_map = compute_persetpoint_from_wide_report(path, target_c, (ref_start, ref_end), first_serial_col)
        for serial, stats_dict in sp_map.items():
            entry = {"target_c": target_c, **stats_dict}
            all_serial_sp.setdefault(serial, []).append(entry)

    return all_serial_sp


def emit_batch(batch_name: str, cert_date_str: str, test_jp: str,
               run_ts: str, serial_sp: dict[str, list], setpoints_json: list):
    run_id = str(uuid.uuid4())
    all_serials = sorted(serial_sp.keys())
    if not all_serials:
        print(f"-- WARN: no serial data for {batch_name}", file=sys.stderr)
        return

    print(f"\n-- === {batch_name} ({len(all_serials)} loggers) ===")
    print(
        f"INSERT INTO calibration_runs "
        f"(id, batch_name, status, testing_start, testing_end, certificate_date, "
        f"threshold_c, setpoints, start_cert_no, cert_width, test_date_jp, doc_date_jp, "
        f"created_at, completed_at) VALUES ("
        f"{q(run_id)}, {q(batch_name)}, 'complete', "
        f"{q(run_ts)}, {q(run_ts)}, {q(cert_date_str)}, "
        f"0.5, {qj(setpoints_json)}, '', 10, {q(test_jp)}, {q(test_jp)}, "
        f"{q(run_ts)}, {q(run_ts)}) ON CONFLICT DO NOTHING;"
    )

    for serial in all_serials:
        sp_list = serial_sp[serial]

        # Verdict: pass only if all setpoints within tolerance
        verdict = "pass" if all(sp["within_tol"] for sp in sp_list) else "fail"

        # max_deviation_c: largest absolute deviation across setpoints
        max_dev = max(abs(sp["dev_c"]) for sp in sp_list)

        lg_id = str(uuid.uuid4())
        res_id = str(uuid.uuid4())

        print(
            f"INSERT INTO loggers (id, serial_no, created_at) VALUES "
            f"({q(lg_id)}, {q(serial)}, {q(run_ts)}) ON CONFLICT (serial_no) DO NOTHING;"
        )
        print(
            f"INSERT INTO logger_results "
            f"(id, run_id, logger_id, sheet_name, verdict, per_setpoint, max_deviation_c, created_at) "
            f"SELECT {q(res_id)}, {q(run_id)}, id, {q(serial)}, {q(verdict)}, "
            f"{qj(sp_list)}, {round(max_dev, 3)}, {q(run_ts)} "
            f"FROM loggers WHERE serial_no = {q(serial)} ON CONFLICT DO NOTHING;"
        )


print("BEGIN;")

# October 2025
oct_sp = build_oct_persetpoint()
oct_setpoints = [
    {"target_c": 40,  "start_at": "2025-10-28T15:00:00+00:00", "end_at": "2025-10-28T18:00:00+00:00"},
    {"target_c": 5,   "start_at": "2025-10-29T06:00:00+00:00", "end_at": "2025-10-29T12:00:00+00:00"},
    {"target_c": -40, "start_at": "2025-10-29T13:00:00+00:00", "end_at": "2025-10-29T16:00:00+00:00"},
]
emit_batch(
    batch_name="Calibration October 2025 — ITE Calibration",
    cert_date_str="2025-10-29",
    test_jp="2025年10月29日",
    run_ts="2025-10-29 09:00:00+00",
    serial_sp=oct_sp,
    setpoints_json=oct_setpoints,
)

# November 2025
nov_sp = build_nov_persetpoint()
nov_setpoints = [
    {"target_c": -40, "start_at": "2025-11-05T17:00:00+00:00", "end_at": "2025-11-06T17:00:00+00:00"},
    {"target_c": 5,   "start_at": "2025-11-07T00:00:00+00:00", "end_at": "2025-11-07T14:00:00+00:00"},
    {"target_c": 40,  "start_at": "2025-11-14T00:00:00+00:00", "end_at": "2025-11-14T02:00:00+00:00"},
]
emit_batch(
    batch_name="Calibration November 2025 — ITE Calibration",
    cert_date_str="2025-11-14",
    test_jp="2025年11月14日",
    run_ts="2025-11-14 09:00:00+00",
    serial_sp=nov_sp,
    setpoints_json=nov_setpoints,
)

print("\nCOMMIT;")
