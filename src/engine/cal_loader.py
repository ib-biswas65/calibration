"""Calibration logger XLSX loading — preserves original sheet parsing logic."""

from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


def load_workbook(path: Path):
    """Load the calibration XLSX workbook (data_only=True) and return (wb, sheet_names)."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    return wb, wb.sheetnames


def load_calibration_sheet(wb, sheet_name: str) -> pd.DataFrame:
    """
    Load one calibration logger sheet into a DataFrame.
    Columns expected: #, DateTime, Temp1, Temp2, Hum1, Hum2, Light, Vibration, Battery
    We use column index [1] for datetime and [3] for Temp2.
    """
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1] is None:
            continue
        ts_str = str(row[1]).strip()
        try:
            ts = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
        try:
            temp2 = float(str(row[3]).strip())
        except (ValueError, TypeError):
            continue
        rows.append((ts, temp2))
    return pd.DataFrame(rows, columns=["timestamp", "temp"])
